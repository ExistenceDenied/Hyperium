from __future__ import annotations

from pathlib import Path

from core.tools.tool import Tool
from infrastructure.tools.scoped import confine

_MAX_ROWS = 500


class ReadExcelTool(Tool):
    """Read a sheet from an .xlsx file, confined to a root directory."""

    name = "read_excel"
    description = (
        "Read a sheet from an .xlsx spreadsheet and return its cells as text. "
        "Use this before updating a spreadsheet, to see what is already there."
    )
    parameters = {
        "type": "object",
        "properties": {
            "path": {"type": "string", "description": "Path to the .xlsx file."},
            "sheet": {
                "type": "string",
                "description": "Sheet name; leave blank for the first sheet.",
            },
        },
        "required": ["path"],
    }

    def __init__(self, root: Path) -> None:
        self._root = root.resolve()

    def invoke(self, arguments: dict) -> str:
        raw = str(arguments.get("path", "")).strip()
        target = confine(self._root, raw) if raw else None

        if target is None:
            return f"Error: '{raw}' is outside the permitted directory."
        if not target.is_file():
            return (
                f"Error: '{raw}' does not exist yet. To create a new "
                "spreadsheet, use write_excel instead of reading it first."
            )

        try:
            from openpyxl import load_workbook

            workbook = load_workbook(target, data_only=True)
        except Exception as error:
            return f"Error: could not open '{raw}': {error}"

        sheet = str(arguments.get("sheet", "")).strip()
        chosen = sheet if sheet and sheet in workbook.sheetnames else None
        worksheet = workbook[chosen] if chosen else workbook.active

        rows = []
        for row in worksheet.iter_rows(values_only=True):
            rows.append(" | ".join("" if cell is None else str(cell) for cell in row))
            if len(rows) >= _MAX_ROWS:
                break

        body = "\n".join(rows) or "(empty)"
        return f"Sheet '{worksheet.title}':\n{body}"


class WriteExcelTool(Tool):
    """Create or overwrite a sheet with rows of data, confined to a root."""

    name = "write_excel"
    description = (
        "Create or replace a sheet in an .xlsx file with rows of data. The "
        "first row is the header. Creates the file if it does not exist. Use "
        "this to produce a spreadsheet such as an invoice, quote or tracker."
    )
    parameters = {
        "type": "object",
        "properties": {
            "path": {"type": "string", "description": "Path to the .xlsx file."},
            "sheet": {"type": "string", "description": "Sheet name (default Sheet1)."},
            "rows": {
                "type": "array",
                "items": {"type": "array", "items": {"type": "string"}},
                "description": "Rows of cells; the first row is the header.",
            },
        },
        "required": ["path", "rows"],
    }

    def __init__(self, root: Path) -> None:
        self._root = root.resolve()

    def preview(self, arguments: dict) -> str:
        raw = str(arguments.get("path", "")).strip()
        rows = arguments.get("rows") or []
        sheet = str(arguments.get("sheet", "Sheet1")).strip() or "Sheet1"
        return f"Write {len(rows)} rows to sheet '{sheet}' in {raw}."

    def invoke(self, arguments: dict) -> str:
        raw = str(arguments.get("path", "")).strip()
        target = confine(self._root, raw) if raw else None

        if target is None:
            return f"Error: '{raw}' is outside the permitted directory."

        rows = arguments.get("rows") or []
        sheet = str(arguments.get("sheet", "Sheet1")).strip() or "Sheet1"

        try:
            from openpyxl import Workbook, load_workbook

            workbook = load_workbook(target) if target.exists() else Workbook()

            if sheet in workbook.sheetnames:
                worksheet = workbook[sheet]
                worksheet.delete_rows(1, worksheet.max_row)
            elif _is_blank_default(workbook):
                worksheet = workbook.active
                worksheet.title = sheet
            else:
                worksheet = workbook.create_sheet(sheet)

            # Explicit cell writes, not append: append can leave a phantom
            # blank first row on a freshly created sheet.
            for r, row in enumerate(rows, start=1):
                for c, value in enumerate(row, start=1):
                    worksheet.cell(row=r, column=c, value=str(value))

            target.parent.mkdir(parents=True, exist_ok=True)
            workbook.save(target)
        except Exception as error:
            return f"Error: could not write '{raw}': {error}"

        return f"Wrote {len(rows)} rows to sheet '{sheet}' in {raw}."


class UpdateExcelCellTool(Tool):
    """Set a single cell in an existing .xlsx file, confined to a root."""

    name = "update_excel_cell"
    description = (
        "Set the value of one cell in an existing .xlsx file — for example set "
        "B3 to a new amount. Read the sheet first to find the right cell."
    )
    parameters = {
        "type": "object",
        "properties": {
            "path": {"type": "string", "description": "Path to the .xlsx file."},
            "sheet": {"type": "string", "description": "Sheet name; blank for first."},
            "cell": {"type": "string", "description": "Cell reference, e.g. B3."},
            "value": {"type": "string", "description": "The new value."},
        },
        "required": ["path", "cell", "value"],
    }

    def __init__(self, root: Path) -> None:
        self._root = root.resolve()

    def preview(self, arguments: dict) -> str:
        return (
            f"Set cell {arguments.get('cell')} to '{arguments.get('value')}' "
            f"in {arguments.get('path')}."
        )

    def invoke(self, arguments: dict) -> str:
        raw = str(arguments.get("path", "")).strip()
        target = confine(self._root, raw) if raw else None

        if target is None:
            return f"Error: '{raw}' is outside the permitted directory."
        if not target.is_file():
            return f"Error: '{raw}' does not exist; create it with write_excel first."

        try:
            from openpyxl import load_workbook

            workbook = load_workbook(target)
            sheet = str(arguments.get("sheet", "")).strip()
            worksheet = (
                workbook[sheet]
                if sheet and sheet in workbook.sheetnames
                else workbook.active
            )
            worksheet[str(arguments.get("cell", "A1")).strip()] = str(
                arguments.get("value", "")
            )
            workbook.save(target)
        except Exception as error:
            return f"Error: could not update '{raw}': {error}"

        return f"Set {arguments.get('cell')} to '{arguments.get('value')}' in {raw}."


def _is_blank_default(workbook) -> bool:
    """A freshly created workbook has one empty default sheet we can rename."""
    if len(workbook.sheetnames) != 1:
        return False

    sheet = workbook.active
    return (
        sheet.title in ("Sheet", "Sheet1")
        and sheet.max_row == 1
        and sheet.max_column == 1
        and sheet["A1"].value is None
    )
